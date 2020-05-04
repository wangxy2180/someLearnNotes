import openpyxl
import os


def input_len_check(data_name, data_value, true_len):
    # 只能做纯数字数据的检查
    if len(data_value) != eval(true_len):
        print('error!' + data_name + '应为' + true_len + '位')
        return False
    else:
        if not data_value.isdigit():
            print('error!' + data_name + '应为纯数字')
            return False
        else:
            return True


def add_info(workbook_dict):
    print('----下面开始添加学生信息----')
    new_info = input_info()
    write_temp = []
    # 查找当前excel表中有哪些项，并将对应的项按序放入list，然后append
    for file_name, wb_and_ws in workbook_dict.items():
        for i in range(1, wb_and_ws[1].max_column + 1):
            write_temp.append(new_info[wb_and_ws[1].cell(row=1, column=i).value])
        workbook_dict[file_name][1].append(write_temp)
        write_temp.clear()
        wb_and_ws[0].save(file_name)
        print('文件' + file_name + '新增成功')
    print('----添加成功----')


def search_info(workbook_dict):
    search_stu_name = input('请输入学生姓名以进行查询:')
    # 查找姓名和身份证号在addrInfo中所在列的序号
    name_column_in_addr = search_col_by_data_name(workbook_dict['addrInfo.xlsx'][1], ['姓名'])
    person_id_column_in_addr = search_col_by_data_name(workbook_dict['addrInfo.xlsx'][1], ['身份证号'])
    # 遍历姓名列，查找输入的姓名是否存在
    for i in range(2, workbook_dict['addrInfo.xlsx'][1].max_row + 1):
        if workbook_dict['addrInfo.xlsx'][1].cell(row=i, column=name_column_in_addr).value != search_stu_name:
            if i == workbook_dict['addrInfo.xlsx'][1].max_row:
                print('error！户籍表中查无此人')
            continue
        else:
            print('----查到了！他的信息是：----')
            searched_key_info_dict = {}
            searched_name_row_in_addr = i
            searched_person_id \
                = workbook_dict['addrInfo.xlsx'][1].cell(searched_name_row_in_addr, person_id_column_in_addr).value
            # 暂时认为姓名和身份证号都是唯一的，凭借任意一个就可以确定一个人
            searched_key_info_dict['姓名'] = search_stu_name
            searched_key_info_dict['身份证号'] = searched_person_id
            modify_dict, searched_row_dict = get_info(workbook_dict, searched_key_info_dict)
            show_info_single(modify_dict)
            while 1:
                search_type = eval(input('====请输入序号进行操作====\n1.修改信息\n2.删除该学生\n3.退出\n您的输入是：'))
                if search_type == 1:
                    modify_info(workbook_dict, searched_row_dict, modify_dict)
                    break
                elif search_type == 2:
                    delete_info(workbook_dict, searched_row_dict)
                    break
                elif search_type == 3:
                    break
                else:
                    print('error！输入有误，请重新输入')
            break


def search_col_by_data_name(ws, data_name):
    ret = -1
    for i in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=i).value not in data_name:
            if ws.max_column == i:
                print('没有找到data_name:', data_name)
            continue
        elif ws.cell(row=1, column=i).value in data_name:
            ret = i
            break
        else:
            print('search_col_by_data_name 出现了一些意外')
    return ret


def search_row_in_col_by_data_value(ws, col, data_value):
    ret = -1
    for i in range(1, ws.max_row + 1):
        if ws.cell(row=i, column=col).value not in data_value:
            if ws.max_row == i:
                print('没有找到这个data_value：', data_value)
            continue
        elif ws.cell(row=i, column=col).value in data_value:
            ret = i
            break
        else:
            print('search_row_in_col_by_data_value 出现了一些问题')
    return ret


def delete_info(workbook_dict_del, searched_row_dict_del):
    for file_name, wb_and_ws in workbook_dict_del.items():
        # wb_and_ws = [workbook,worksheet]
        wb_and_ws[1].delete_rows(searched_row_dict_del[file_name])
        wb_and_ws[0].save(file_name)
        print(file_name + '中删除完成')


def modify_info(workbook_dict_modify, searched_row_dict_modify, modify_dict):
    modified_dict = {}
    while 1:
        modify_data_name = input('====请输入要修改的学生信息名称,修改完毕请输入0==== \n您的输入是：')
        if modify_data_name == '0':
            print('----修改完成！----')
            break
        if modify_data_name not in modify_dict:
            print('error！你输入的修改项不存在，请重新输入')
            continue
        modify_data = input('===请输入修改后的值===\n')
        modified_dict[modify_data_name] = modify_data
    #     将修改后的数据写回所有的表格
    for file_name, wb_and_ws in workbook_dict_modify.items():
        for j in range(1, wb_and_ws[1].max_column + 1):
            data_name_in_xlsx = wb_and_ws[1].cell(1, j).value
            if data_name_in_xlsx in modified_dict:
                wb_and_ws[1].cell(row=searched_row_dict_modify[file_name], column=j).value \
                    = modified_dict[data_name_in_xlsx]
        wb_and_ws[0].save(file_name)
    return modify_dict


def get_info(workbook_dict_show, searched_key_info_show):
    # 通过key_info获取这个人的所有信息
    show_dict = {}
    searched_row_dict_ret = {}
    for file_name, wb_and_ws in workbook_dict_show.items():
        # 查找当前文件中key_info所在的列，然后再根据key_info的值查找所在的行，并将其放入字典
        key_info_col = search_col_by_data_name(wb_and_ws[1], searched_key_info_show.keys())
        searched_row = search_row_in_col_by_data_value(wb_and_ws[1], key_info_col, searched_key_info_show.values())
        searched_row_dict_ret[file_name] = searched_row
        # 根据key_info的数据名，将key_info放入字典
        for j in range(1, wb_and_ws[1].max_column + 1):
            show_dict[wb_and_ws[1].cell(1, j).value] \
                = wb_and_ws[1].cell(row=searched_row, column=j).value
    # for key, value in show_dict.items():
    #     print(key, ':', value)
    return show_dict, searched_row_dict_ret


def show_info_single(print_dict):
    for key, value in print_dict.items():
        print(key, ':', value)


def show_info_all(workbook_dict_show_all):
    key_info_dict = {}
    for file_name, wb_and_ws in workbook_dict_show_all.items():
        # 在户籍表中查找姓名和身份证号所在的列
        if file_name == 'addrInfo.xlsx':
            name_col = search_col_by_data_name(wb_and_ws[1], '姓名')
            person_id_col = search_col_by_data_name(wb_and_ws[1], '身份证号')
            for i in range(2, wb_and_ws[1].max_row + 1):
                # 加入key_info字典
                key_info_dict['姓名'] = wb_and_ws[1].cell(i, name_col).value
                key_info_dict['身份证号'] = wb_and_ws[1].cell(i, person_id_col).value
                # 通过key_info获取这个人的所有信息
                show_dict_all, unuse_value = get_info(workbook_dict_show_all, key_info_dict)
                print('姓名:{name}; 学号:{stu_id}; 身份证号:{id}; 出生日期:{bir}; 学院:{coll}; 专业:{major}; 班级:{cla}; 地址:{addr}; '
                      '电话:{tel}; 微信:{wech}; 邮件:{email}'.format(
                        name=show_dict_all['姓名'],
                        id=show_dict_all['身份证号'],
                        bir=show_dict_all['出生日期'],
                        addr=show_dict_all['地址'],
                        stu_id=show_dict_all['学号'],
                        coll=show_dict_all['学院'],
                        major=show_dict_all['专业'],
                        cla=show_dict_all['班级'],
                        tel=show_dict_all['电话'],
                        wech=show_dict_all['微信'],
                        email=show_dict_all['邮件'])
                      )
    print('----全部人员打印完成----')


def input_info():
    name = input('请输入学生姓名：')
    while 1:
        person_id = input('请输入身份证号：')
        if input_len_check('身份证号', person_id, '18'):
            break
    while 1:
        stu_id = input('请输入学号（9位）：')
        if input_len_check('学号', stu_id, '9'):
            break
    while 1:
        birth = input('请输入出生日期（8位）：')
        if input_len_check('学号', birth, '8'):
            break
    while 1:
        tel = input('请输入电话（11位）：')
        if input_len_check('学号', tel, '11'):
            break
    new_info = {'姓名': name,
                '身份证号': person_id,
                '学号': stu_id,
                '学院': input('请输入学院:'),
                '专业': input('请输入专业:'),
                '班级': input('请输入班级:'),
                '出生日期': birth,
                '地址': input('请输入地址:'),
                '电话': tel,
                '微信': input('请输入微信'),
                '邮件': input('请输入邮件')
                }
    return new_info


def load_xlsx():
    # 导入当前目录所有xlsx文件，存入一个字典中
    workbook_dict = {}
    for file in os.listdir(os.getcwd()):
        if os.path.splitext(file)[1] == '.xlsx':
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            workbook_dict[file] = [wb, ws]
    print('发现文件', end="")
    print(workbook_dict.keys())
    return workbook_dict


def main():
    workbook_dict = load_xlsx()
    while 1:
        op_code = input('====请输入序号进行操作====\n1.添加学生信息\n2.查询学生信息\n3.打印所有学生资料\n4.退出程序\n您的输入是：')
        if op_code == '1':
            add_info(workbook_dict)
        elif op_code == '2':
            search_info(workbook_dict)
        elif op_code == '3':
            show_info_all(workbook_dict)
        elif op_code == '4':
            break
        else:
            print('error！您的输入有误，请重新输入')
    print('程序结束，再见！')


main()
